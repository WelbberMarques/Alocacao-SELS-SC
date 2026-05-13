
import re
from openpyxl import load_workbook
from openpyxl.styles import Font

VERMELHO = "FFFF0000"
PRETO    = "FF000000"


def atualizar_consulta1(caminho: str, log_fn=None) -> bool:
    """
    Abre a planilha no Excel em background via xlwings,
    forca refresh de todas as conexoes/queries e salva.
    Retorna True se conseguiu atualizar.
    """
    def log(msg, tipo="info"):
        if log_fn:
            log_fn(msg, tipo)

    try:
        import xlwings as xw
        log("   Abrindo Excel para atualizar Consulta1...", "info")

        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        app.screen_updating = False

        try:
            wb = app.books.open(caminho)

            import time
            try:
                # Tenta atualizar cada conexao individualmente
                conexoes = wb.api.Connections
                total_conn = conexoes.Count
                log(f"   {total_conn} conexao(oes) encontrada(s)", "info")

                for i in range(1, total_conn + 1):
                    try:
                        conn = conexoes.Item(i)
                        nome = conn.Name
                        log(f"   Atualizando: {nome}...", "info")
                        conn.Refresh()
                        time.sleep(1)
                    except Exception as ec:
                        log(f"   Conexao {i}: {ec}", "warn")

                # Tenta tambem via queries do Power Query
                try:
                    for q in wb.api.Queries:
                        try:
                            q.Refresh()
                            time.sleep(0.5)
                        except Exception:
                            pass
                except Exception:
                    pass

                # Fallback: RefreshAll
                wb.api.RefreshAll()
                time.sleep(3)

                # Forca calculo
                wb.api.Application.Calculate()
                time.sleep(1)

                log("   Consulta1 atualizada!", "ok")

            except Exception as e:
                log(f"   Erro ao atualizar conexoes: {e}", "warn")
                try:
                    wb.api.RefreshAll()
                    time.sleep(3)
                except Exception:
                    pass

            wb.save()
            wb.close()
            return True

        finally:
            app.quit()

    except ImportError:
        log("   xlwings nao instalado — usando cache da Consulta1", "warn")
        return False
    except Exception as e:
        log(f"   Erro ao atualizar Consulta1: {e}", "warn")
        return False


def _atualizar_valor_unit(ws, celula, novo_valor, log_fn):
    cell = ws[celula]
    valor_atual = cell.value or 0.0
    mudou = abs(float(valor_atual) - float(novo_valor)) > 0.001
    cell.value = novo_valor
    fonte = cell.font
    cell.font = Font(
        name=fonte.name, size=fonte.size, bold=fonte.bold,
        italic=fonte.italic, color=VERMELHO if mudou else PRETO,
    )
    if mudou:
        log_fn(f"   ! {celula}: {valor_atual} -> {novo_valor} (alterado)", "erro")
    else:
        log_fn(f"   + {celula}: {novo_valor}", "ok")


def _ler_consulta1(wb, log_fn):
    """
    Le a aba Consulta1 da planilha e retorna
    {numero_resumo (int): valor_total (float)}.
    Vendas + Envio do mesmo resumo sao somados automaticamente.
    """
    aba = None
    for nome in wb.sheetnames:
        if 'consulta' in nome.lower():
            aba = wb[nome]
            break

    if aba is None:
        log_fn("   Aba Consulta1 nao encontrada na planilha", "warn")
        return {}

    totais = {}
    for row in aba.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 5:
            continue
        valor  = row[1]  # coluna B: Multiplicacao
        nf_raw = row[4]  # coluna E: numero do resumo extraido
        if valor is None or nf_raw is None:
            continue
        try:
            nf  = int(float(str(nf_raw)))
            val = float(valor)
            totais[nf] = totais.get(nf, 0.0) + val
        except (ValueError, TypeError):
            pass

    log_fn(f"   Consulta1: {len(totais)} resumo(s) encontrado(s)", "info")
    return totais


def _comparar_coluna_t(ws, totais_consulta, log_fn):
    """
    Compara coluna T com Consulta1.
    Regras:
    - So analisa linhas que tem numero de resumo na coluna A
    - Se o resumo NAO esta na Consulta1: deixa como esta, nao pinta
    - Se o resumo ESTA na Consulta1 e valor DIVERGE: atualiza e pinta VERMELHO
    - Se o resumo ESTA na Consulta1 e valor BATE: garante cor PRETA (sem alteracao)
    """
    divergencias = 0
    soma_t = 0.0

    for linha in range(2, 27):
        val_a = ws.cell(row=linha, column=1).value
        if val_a is None or str(val_a).strip() == "":
            continue

        m = re.match(r'(\d+)', str(val_a).strip())
        if not m:
            continue
        try:
            nf_num = int(m.group(1))
        except (ValueError, TypeError):
            continue

        # Se nao esta na Consulta1 — deixa exatamente como esta, sem tocar
        if nf_num not in totais_consulta:
            log_fn(f"   - T{linha} (NF {nf_num}): nao esta na Consulta1, mantido", "info")
            continue

        total  = round(totais_consulta[nf_num], 2)
        soma_t += total

        cell_t = ws.cell(row=linha, column=20)
        try:
            atual = round(float(cell_t.value or 0), 2)
        except (ValueError, TypeError):
            atual = 0.0

        f = cell_t.font
        # Sempre atualiza valor e pinta VERMELHO se esta na Consulta1
        if abs(atual - total) > 0.01:
            cell_t.value = total
            log_fn(f"   ! T{linha} (NF {nf_num}): {atual} -> {total}", "erro")
            divergencias += 1
        else:
            log_fn(f"   + T{linha} (NF {nf_num}): {total} ok", "ok")
        # Sempre vermelho — esta na Consulta1
        cell_t.font = Font(name=f.name, size=f.size,
                           bold=f.bold, italic=f.italic,
                           color=VERMELHO)

    # T28: total geral
    if soma_t > 0:
        cell_t28 = ws.cell(row=28, column=20)
        try:
            atual_28 = round(float(cell_t28.value or 0), 2)
        except (ValueError, TypeError):
            atual_28 = 0.0
        if abs(atual_28 - round(soma_t, 2)) > 0.01:
            cell_t28.value = round(soma_t, 2)
            f = cell_t28.font
            cell_t28.font = Font(name=f.name, size=f.size,
                                 bold=f.bold, italic=f.italic, color=VERMELHO)
            log_fn(f"   ! T28 total: {atual_28} -> {round(soma_t, 2)}", "erro")
            divergencias += 1

    return divergencias


def preencher_planilha(caminho_arquivo, dados, meses, log_fn,
                       caminho_modelo=None, totais_consulta=None):
    """
    Abre caminho_arquivo, cria/atualiza a aba com numero da NF,
    preenche com dados do GSA, e compara coluna T com Consulta1.
    totais_consulta: dict ja carregado {nf: valor} - preferido
    caminho_modelo: caminho do .xlsm para ler Consulta1 se totais nao fornecido
    """
    numero_nf  = dados["numero_nf"]
    valor_vs   = dados["valor_vs"]
    valor_na   = dados["valor_na"]
    valor_najr = dados["valor_najr"]

    # Carrega sem keep_vba para evitar corrompimento das conexoes externas
    # A planilha preenchida e sempre salva como .xlsx
    wb = load_workbook(caminho_arquivo)

    # Cria nova aba copiando do Modelo, ou usa aba existente
    if str(numero_nf) in wb.sheetnames:
        ws = wb[str(numero_nf)]
        log_fn(f"   > Aba existente: {numero_nf}", "info")
    elif "Modelo " in wb.sheetnames or "Modelo" in wb.sheetnames:
        nome_modelo = "Modelo " if "Modelo " in wb.sheetnames else "Modelo"
        ws = wb.copy_worksheet(wb[nome_modelo])
        ws.title = str(numero_nf)
        log_fn(f"   > Nova aba criada: {numero_nf}", "info")
    else:
        ws = wb.active
        log_fn(f"   > Aba ativa: {ws.title}", "info")

    # R1: numero da NF
    ws["R1"] = int(numero_nf)

    # S28 tem formula no modelo (=(R27-R28)-R29) - nao sobrescrever

    # C1, E1, G1: valores unitarios
    log_fn("   Valores unitarios...", "info")
    _atualizar_valor_unit(ws, "C1", valor_vs,   log_fn)
    _atualizar_valor_unit(ws, "E1", valor_na,   log_fn)
    _atualizar_valor_unit(ws, "G1", valor_najr, log_fn)

    # Linhas 2-26: resumos do GSA
    for i, num_resumo in enumerate(sorted(meses.keys())):
        linha = i + 2
        if linha > 26:
            log_fn(f"   ! Resumo {num_resumo} excede limite de 25 linhas", "erro")
            break

        info      = meses[num_resumo]
        cancelado = info.get("cancelado", False)
        qtd_vs    = info.get("vs",   0)
        qtd_na    = info.get("na",   0)
        qtd_najr  = info.get("najr", 0)

        valor_a = f"{num_resumo} - CANCELADO" if cancelado else num_resumo
        cell_a  = ws.cell(row=linha, column=1)
        cell_a.value = valor_a

        if cancelado:
            cell_a.font = Font(color=VERMELHO, bold=True)
            for col in [2, 3, 4, 5, 6, 7]:
                ws.cell(row=linha, column=col).font = Font(color=VERMELHO)

        if qtd_vs   > 0: ws.cell(row=linha, column=2).value = qtd_vs
        if qtd_na   > 0: ws.cell(row=linha, column=4).value = qtd_na
        if qtd_najr > 0: ws.cell(row=linha, column=6).value = qtd_najr

        log_fn(
            f"   Linha {linha} - {valor_a}: VS={qtd_vs} NA={qtd_na} NAJR={qtd_najr}",
            "erro" if cancelado else "ok"
        )

    # Linha 28: quantidades do XML
    ws["A28"] = int(numero_nf)
    ws["B28"] = dados["vs"]
    ws["D28"] = dados["na"]
    ws["F28"] = dados["najr"]

    # Compara coluna T com Consulta1
    log_fn("   Comparando coluna T com Consulta1...", "info")

    # Usa totais ja carregados (preferido - evita reler o arquivo a cada NF)
    if not totais_consulta and caminho_modelo:
        try:
            keep = caminho_modelo.lower().endswith('.xlsm')
            wb_modelo = load_workbook(caminho_modelo, keep_vba=keep, data_only=True)
            totais_consulta = _ler_consulta1(wb_modelo, log_fn)
        except Exception as e_modelo:
            log_fn(f"   Erro ao ler modelo: {e_modelo}", "warn")
            totais_consulta = _ler_consulta1(wb, log_fn)
    elif not totais_consulta:
        totais_consulta = _ler_consulta1(wb, log_fn)

    if totais_consulta:
        divs = _comparar_coluna_t(ws, totais_consulta, log_fn)
        if divs:
            log_fn(f"   ! {divs} divergencia(s) corrigida(s) na coluna T", "warn")
        else:
            log_fn("   + Coluna T OK, sem divergencias", "ok")
    else:
        log_fn("   Consulta1 nao encontrada, comparacao ignorada", "warn")

    wb.save(caminho_arquivo)
    log_fn(f"   + Aba {numero_nf} salva", "ok")
    return caminho_arquivo
